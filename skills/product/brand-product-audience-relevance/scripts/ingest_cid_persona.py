#!/usr/bin/env python3
import argparse
import csv
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path("/Users/apple/Downloads/User/CID 人群")
OUT_DIR = Path("/Users/apple/Documents/Codex/2026-05-22/2026/outputs/cid_persona_design")
LOAD_DIR = OUT_DIR / "load"
SCHEMA_SQL = OUT_DIR / "cid_persona_schema.sql"
LOAD_SQL = OUT_DIR / "load_cid_persona.sql"


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def text(value):
    value = clean(value)
    return str(value) if value is not None else None


def metric_value(value, default=0):
    return value if isinstance(value, (int, float)) else default


def nullable_metric(value):
    return value if isinstance(value, (int, float)) else None


def dt_from_file(date_s, time_s):
    try:
        return datetime.strptime(date_s + time_s, "%Y%m%d%H%M%S").isoformat(sep=" ")
    except Exception:
        return None


def parse_file(path):
    name = path.name
    m = re.match(r"(.+?)x(.+?) Persona Report (\d{8}) (\d{6})\.xlsx$", name)
    if not m:
        raise ValueError(f"Unexpected file name: {name}")
    lifecycle, city_tier, date_s, time_s = m.groups()

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["人群画像"]
    rows = [tuple(clean(v) for v in row) for row in ws.iter_rows(values_only=True)]

    segment_name = text(rows[4][5]) if len(rows) > 4 and len(rows[4]) > 5 else f"{lifecycle}x{city_tier}"
    segment_definition = text(rows[5][5]) if len(rows) > 5 and len(rows[5]) > 5 else None

    ecom_header = None
    app_header = None
    for idx, row in enumerate(rows, start=1):
        joined = "|".join(str(v) for v in row if v is not None)
        if ecom_header is None and ("购买频次" in joined or "购买金额" in joined):
            ecom_header = idx
        if app_header is None and ("日均使用频次" in joined or "日均使用时长" in joined):
            app_header = idx

    if ecom_header is None:
        raise ValueError(f"No purchase header found: {name}")

    total = None
    for row in rows[7 : ecom_header - 1]:
        vals = list(row) + [None] * 8
        one_id, ratio = vals[5], vals[6]
        if isinstance(one_id, (int, float)) and isinstance(ratio, (int, float)) and abs(ratio - 1) < 1e-9:
            total = max(total or 0, int(one_id))

    report = {
        "file_name": name,
        "file_path": str(path),
        "sheet_name": "人群画像",
        "report_generated_at": dt_from_file(date_s, time_s),
    }
    segment = {
        "segment_code": f"{lifecycle}x{city_tier}",
        "segment_name": segment_name,
        "lifecycle_stage": lifecycle,
        "city_tier": city_tier,
        "segment_definition": segment_definition,
        "segment_total_one_id": total,
        "file_name": name,
    }

    profile_rows = parse_profile_rows(rows, ecom_header, segment["segment_code"])
    entity_rows, bucket_rows = parse_purchase_rows(rows, ecom_header, app_header, segment["segment_code"])
    app_rows = parse_media_rows(rows, app_header, segment["segment_code"]) if app_header else []
    entity_rows.extend(app_rows)

    return report, segment, profile_rows, entity_rows, bucket_rows


def parse_profile_rows(rows, ecom_header, segment_code):
    profile_rows = []
    level1 = None
    level2 = None
    level2_override = None

    for source_row, row in enumerate(rows[7 : ecom_header - 1], start=8):
        vals = list(row) + [None] * 21
        c2, c3, c4, c5, c6, c7, c8 = vals[1], vals[2], vals[3], vals[4], vals[5], vals[6], vals[7]

        if c2:
            level1 = text(c2)
            level2 = None
            level2_override = None
        if c3:
            level2 = text(c3)
            level2_override = None
        if c4:
            level2_override = text(c4)

        level3 = text(c5)
        if not level1 or not level3:
            continue

        resolved_level2 = level2_override or level2 or level1
        has_one_id = isinstance(c6, (int, float))
        profile_rows.append({
            "segment_code": segment_code,
            "level1_category": level1,
            "level2_category": resolved_level2,
            "level3_value": level3,
            "one_id": int(c6) if has_one_id else 0,
            "ratio": c7 if isinstance(c7, (int, float)) else 0,
            "tgi": c8 if isinstance(c8, (int, float)) else None,
            "is_zero_row": not has_one_id,
            "source_row": source_row,
        })

    return profile_rows


def parse_bucket_headers(rows, ecom_header):
    bucket_headers = []
    if ecom_header and ecom_header + 1 <= len(rows):
        header = list(rows[ecom_header]) + [None] * 21
        for col_idx in range(9, 22):
            label = text(header[col_idx - 1])
            if not label:
                continue
            family = "purchase_frequency_online" if col_idx <= 12 else "purchase_amount_online"
            bucket_headers.append((col_idx, family, label, col_idx))
    return bucket_headers


def parse_purchase_rows(rows, ecom_header, app_header, segment_code):
    entity_rows = []
    bucket_rows = []
    bucket_headers = parse_bucket_headers(rows, ecom_header)
    cur_category = None
    ecom_end = (app_header - 1) if app_header else len(rows)

    for source_row, row in enumerate(rows[ecom_header + 1 : ecom_end], start=ecom_header + 2):
        vals = list(row) + [None] * 21
        c2, c5, c6, c7, c8 = vals[1], vals[4], vals[5], vals[6], vals[7]

        if c2 and text(c2) != "标签":
            cur_category = text(c2)

        label = text(c5)
        has_one_id = isinstance(c6, (int, float))

        # Rows like "美妆个护 / 美妆个护" without ONE ID are section markers.
        if not label or not has_one_id:
            continue

        if cur_category == "品类购买":
            level2_category = label
            entity_type = "category"
            taxonomy_name = label
        else:
            level2_category = cur_category
            entity_type = "brand"
            taxonomy_name = cur_category

        label_path = " / ".join(v for v in ["品类购买", level2_category, label] if v)
        entity_rows.append({
            "segment_code": segment_code,
            "domain": "purchase",
            "level1_category": "品类购买",
            "level2_category": level2_category,
            "entity_type": entity_type,
            "entity_name": label,
            "taxonomy_node_type": "purchase_category" if taxonomy_name else None,
            "taxonomy_node_name": taxonomy_name,
            "one_id": int(c6),
            "ratio": c7 if isinstance(c7, (int, float)) else 0,
            "tgi": c8 if isinstance(c8, (int, float)) else None,
            "avg_daily_usage_frequency": None,
            "avg_daily_usage_minutes": None,
            "is_zero_row": False,
            "source_row": source_row,
            "label_path": label_path,
            "raw_label": label,
        })

        for col_idx, family, bucket_label, bucket_order in bucket_headers:
            count = metric_value(vals[col_idx - 1], default=0)
            bucket_rows.append({
                "segment_code": segment_code,
                "source_row": source_row,
                "metric_family": family,
                "bucket_label": bucket_label,
                "bucket_order": bucket_order,
                "one_id": int(count),
            })

    return entity_rows, bucket_rows


def parse_media_rows(rows, app_header, segment_code):
    entity_rows = []
    cur_group = None
    cur_category = None

    # Start at the first category marker after the metric header. This preserves 新闻资讯.
    for source_row, row in enumerate(rows[app_header:], start=app_header + 1):
        vals = list(row) + [None] * 10
        c2, c3, c5, c6, c7, c8, c9, c10 = vals[1], vals[2], vals[4], vals[5], vals[6], vals[7], vals[8], vals[9]

        if c2:
            cur_group = text(c2)
        if c3:
            cur_category = text(c3)

        label = text(c5)
        has_one_id = isinstance(c6, (int, float))

        # Rows with category names but no ONE ID are taxonomy markers, not zero app rows.
        if not label or not has_one_id:
            continue

        label_path = " / ".join(v for v in ["媒体行为", cur_category, label] if v)
        entity_rows.append({
            "segment_code": segment_code,
            "domain": "media",
            "level1_category": "媒体行为",
            "level2_category": cur_category,
            "entity_type": "app",
            "entity_name": label,
            "taxonomy_node_type": "media_category" if cur_category else None,
            "taxonomy_node_name": cur_category,
            "one_id": int(c6),
            "ratio": c7 if isinstance(c7, (int, float)) else 0,
            "tgi": c8 if isinstance(c8, (int, float)) else None,
            "avg_daily_usage_frequency": nullable_metric(c9),
            "avg_daily_usage_minutes": nullable_metric(c10),
            "is_zero_row": False,
            "source_row": source_row,
            "label_path": label_path,
            "raw_label": label,
        })

    return entity_rows


def write_csv(path, fieldnames, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: "" if row.get(k) is None else row.get(k) for k in fieldnames})


def q(path):
    return "'" + str(path).replace("'", "''") + "'"


def build_load_sql():
    sql = f"""
\\set ON_ERROR_STOP on
begin;

drop table if exists stg_reports;
drop table if exists stg_segments;
drop table if exists stg_profile;
drop table if exists stg_entities;
drop table if exists stg_buckets;

create temp table stg_reports (
  file_name text,
  file_path text,
  sheet_name text,
  report_generated_at timestamp
);

create temp table stg_segments (
  segment_code text,
  segment_name text,
  lifecycle_stage text,
  city_tier text,
  segment_definition text,
  segment_total_one_id bigint,
  file_name text
);

create temp table stg_profile (
  segment_code text,
  level1_category text,
  level2_category text,
  level3_value text,
  one_id bigint,
  ratio numeric,
  tgi numeric,
  is_zero_row boolean,
  source_row int
);

create temp table stg_entities (
  segment_code text,
  domain text,
  level1_category text,
  level2_category text,
  entity_type text,
  entity_name text,
  taxonomy_node_type text,
  taxonomy_node_name text,
  one_id bigint,
  ratio numeric,
  tgi numeric,
  avg_daily_usage_frequency numeric,
  avg_daily_usage_minutes numeric,
  is_zero_row boolean,
  source_row int,
  label_path text,
  raw_label text
);

create temp table stg_buckets (
  segment_code text,
  source_row int,
  metric_family text,
  bucket_label text,
  bucket_order int,
  one_id bigint
);

\\copy stg_reports from {q(LOAD_DIR / "reports.csv")} csv header
\\copy stg_segments from {q(LOAD_DIR / "segments.csv")} csv header
\\copy stg_profile from {q(LOAD_DIR / "profile.csv")} csv header
\\copy stg_entities from {q(LOAD_DIR / "entities.csv")} csv header
\\copy stg_buckets from {q(LOAD_DIR / "buckets.csv")} csv header

insert into cid.dim_report_source(file_name, file_path, sheet_name, report_generated_at)
select file_name, file_path, sheet_name, report_generated_at from stg_reports
on conflict (file_name) do update set
  file_path = excluded.file_path,
  sheet_name = excluded.sheet_name,
  report_generated_at = excluded.report_generated_at,
  ingested_at = now();

insert into cid.dim_audience_segment(segment_code, segment_name, lifecycle_stage, city_tier, segment_definition, segment_total_one_id, report_id)
select s.segment_code, s.segment_name, s.lifecycle_stage, s.city_tier, s.segment_definition, s.segment_total_one_id, r.report_id
from stg_segments s
join cid.dim_report_source r on r.file_name = s.file_name
on conflict (segment_code) do update set
  segment_name = excluded.segment_name,
  lifecycle_stage = excluded.lifecycle_stage,
  city_tier = excluded.city_tier,
  segment_definition = excluded.segment_definition,
  segment_total_one_id = excluded.segment_total_one_id,
  report_id = excluded.report_id;

insert into cid.fact_segment_profile_value(
  segment_id, level1_category, level2_category, level3_value, one_id, ratio, tgi, is_zero_row, source_row
)
select s.segment_id, p.level1_category, p.level2_category, p.level3_value, p.one_id, p.ratio, p.tgi, p.is_zero_row, p.source_row
from stg_profile p
join cid.dim_audience_segment s on s.segment_code = p.segment_code
on conflict (segment_id, source_row) do update set
  level1_category = excluded.level1_category,
  level2_category = excluded.level2_category,
  level3_value = excluded.level3_value,
  one_id = excluded.one_id,
  ratio = excluded.ratio,
  tgi = excluded.tgi,
  is_zero_row = excluded.is_zero_row;

insert into cid.dim_taxonomy_node(domain, node_type, node_name)
select distinct domain, taxonomy_node_type, taxonomy_node_name
from stg_entities
where taxonomy_node_name is not null and taxonomy_node_name <> ''
on conflict do nothing;

insert into cid.dim_entity(domain, entity_type, entity_name, taxonomy_node_id, canonical_entity_name)
select distinct e.domain, e.entity_type, e.entity_name, t.taxonomy_node_id, e.entity_name
from stg_entities e
left join cid.dim_taxonomy_node t
  on t.domain = e.domain
 and t.node_type = e.taxonomy_node_type
 and t.node_name = e.taxonomy_node_name
on conflict do nothing;

insert into cid.fact_segment_entity_metric(
  segment_id, entity_id, domain, level1_category, level2_category, entity_type, entity_name,
  one_id, ratio, tgi, avg_daily_usage_frequency, avg_daily_usage_minutes, is_zero_row,
  source_row, label_path, raw_label
)
select s.segment_id, de.entity_id, e.domain, e.level1_category, e.level2_category, e.entity_type, e.entity_name,
       e.one_id, e.ratio, e.tgi, e.avg_daily_usage_frequency, e.avg_daily_usage_minutes, e.is_zero_row,
       e.source_row, e.label_path, e.raw_label
from stg_entities e
join cid.dim_audience_segment s on s.segment_code = e.segment_code
left join cid.dim_taxonomy_node t
  on t.domain = e.domain
 and t.node_type = e.taxonomy_node_type
 and t.node_name = e.taxonomy_node_name
join cid.dim_entity de
  on de.domain = e.domain
 and de.entity_type = e.entity_type
 and de.entity_name = e.entity_name
 and coalesce(de.taxonomy_node_id, 0) = coalesce(t.taxonomy_node_id, 0)
on conflict (segment_id, source_row) do update set
  entity_id = excluded.entity_id,
  domain = excluded.domain,
  level1_category = excluded.level1_category,
  level2_category = excluded.level2_category,
  entity_type = excluded.entity_type,
  entity_name = excluded.entity_name,
  one_id = excluded.one_id,
  ratio = excluded.ratio,
  tgi = excluded.tgi,
  avg_daily_usage_frequency = excluded.avg_daily_usage_frequency,
  avg_daily_usage_minutes = excluded.avg_daily_usage_minutes,
  is_zero_row = excluded.is_zero_row,
  label_path = excluded.label_path,
  raw_label = excluded.raw_label;

insert into cid.dim_metric_bucket(metric_family, bucket_label, bucket_order)
select distinct metric_family, bucket_label, bucket_order from stg_buckets
on conflict (metric_family, bucket_label) do update set bucket_order = excluded.bucket_order;

insert into cid.fact_segment_entity_bucket(metric_id, bucket_id, one_id)
select m.metric_id, bkt.bucket_id, b.one_id
from stg_buckets b
join cid.dim_audience_segment s on s.segment_code = b.segment_code
join cid.fact_segment_entity_metric m
  on m.segment_id = s.segment_id
 and m.source_row = b.source_row
join cid.dim_metric_bucket bkt
  on bkt.metric_family = b.metric_family
 and bkt.bucket_label = b.bucket_label
on conflict (metric_id, bucket_id) do update set one_id = excluded.one_id;

commit;
"""
    LOAD_SQL.write_text(sql.strip() + "\n", encoding="utf-8")


def maybe_run_psycopg(args):
    vendor = str(OUT_DIR / "vendor")
    if vendor not in sys.path:
        sys.path.insert(0, vendor)
    import psycopg
    from psycopg import sql

    conninfo = {}
    if args.host:
        conninfo["host"] = args.host
    if args.port:
        conninfo["port"] = int(args.port)
    if args.user:
        conninfo["user"] = args.user
    if args.password:
        conninfo["password"] = args.password
    conninfo["connect_timeout"] = 5

    if args.create_db:
        admin_info = dict(conninfo)
        admin_info["dbname"] = args.admin_dbname
        with psycopg.connect(**admin_info, autocommit=True) as conn:
            exists = conn.execute("select 1 from pg_database where datname = %s", [args.dbname]).fetchone()
            if not exists:
                conn.execute(sql.SQL("create database {}").format(sql.Identifier(args.dbname)))

    db_info = dict(conninfo)
    db_info["dbname"] = args.dbname
    with psycopg.connect(**db_info) as conn:
        with conn.cursor() as cur:
            cur.execute(SCHEMA_SQL.read_text(encoding="utf-8"))
        conn.commit()

    load_text = LOAD_SQL.read_text(encoding="utf-8")
    first_copy = load_text.index("\\copy")
    last_copy_line_end = load_text.rindex("\\copy")
    last_copy_line_end = load_text.index("\n", last_copy_line_end)
    before_copy = "\n".join(
        line for line in load_text[:first_copy].splitlines()
        if not line.startswith("\\set")
    )
    after_copy = load_text[last_copy_line_end + 1 :]

    copy_targets = [
        ("stg_reports", LOAD_DIR / "reports.csv"),
        ("stg_segments", LOAD_DIR / "segments.csv"),
        ("stg_profile", LOAD_DIR / "profile.csv"),
        ("stg_entities", LOAD_DIR / "entities.csv"),
        ("stg_buckets", LOAD_DIR / "buckets.csv"),
    ]

    with psycopg.connect(**db_info) as conn:
        with conn.cursor() as cur:
            cur.execute(before_copy)
            for table, path in copy_targets:
                with path.open("r", encoding="utf-8") as f:
                    with cur.copy(sql.SQL("copy {} from stdin with csv header").format(sql.Identifier(table))) as cp:
                        while True:
                            chunk = f.read(1024 * 1024)
                            if not chunk:
                                break
                            cp.write(chunk)
            cur.execute(after_copy)
        conn.commit()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default=str(BASE_DIR))
    parser.add_argument("--dbname", default="cid_persona")
    parser.add_argument("--host")
    parser.add_argument("--port")
    parser.add_argument("--user")
    parser.add_argument("--password")
    parser.add_argument("--create-db", action="store_true")
    parser.add_argument("--admin-dbname", default="postgres")
    parser.add_argument("--method", choices=["psycopg"], default="psycopg")
    parser.add_argument("--no-run", action="store_true")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    files = sorted(input_dir.glob("*.xlsx"))
    if not files:
        raise SystemExit(f"No xlsx files found in {input_dir}")

    reports, segments, profiles, entities, buckets = [], [], [], [], []
    for path in files:
        report, segment, profile_rows, entity_rows, bucket_rows = parse_file(path)
        reports.append(report)
        segments.append(segment)
        profiles.extend(profile_rows)
        entities.extend(entity_rows)
        buckets.extend(bucket_rows)

    write_csv(LOAD_DIR / "reports.csv", ["file_name", "file_path", "sheet_name", "report_generated_at"], reports)
    write_csv(LOAD_DIR / "segments.csv", ["segment_code", "segment_name", "lifecycle_stage", "city_tier", "segment_definition", "segment_total_one_id", "file_name"], segments)
    write_csv(LOAD_DIR / "profile.csv", ["segment_code", "level1_category", "level2_category", "level3_value", "one_id", "ratio", "tgi", "is_zero_row", "source_row"], profiles)
    write_csv(LOAD_DIR / "entities.csv", ["segment_code", "domain", "level1_category", "level2_category", "entity_type", "entity_name", "taxonomy_node_type", "taxonomy_node_name", "one_id", "ratio", "tgi", "avg_daily_usage_frequency", "avg_daily_usage_minutes", "is_zero_row", "source_row", "label_path", "raw_label"], entities)
    write_csv(LOAD_DIR / "buckets.csv", ["segment_code", "source_row", "metric_family", "bucket_label", "bucket_order", "one_id"], buckets)
    build_load_sql()

    print(f"files={len(files)} reports={len(reports)} segments={len(segments)} profiles={len(profiles)} entities={len(entities)} buckets={len(buckets)}")
    print(f"load_sql={LOAD_SQL}")

    if not args.no_run:
        maybe_run_psycopg(args)
        print("loaded=true")


if __name__ == "__main__":
    main()
