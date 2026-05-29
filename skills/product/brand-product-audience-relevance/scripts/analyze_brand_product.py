from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import math
import re
import statistics
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


AGE_FILE_RE = re.compile(r"(.+?) (一线|新一线|二线|三线|四线五线) Persona Report (\d{8}) (\d{6})\.xlsx$")
LIFE_FILE_RE = re.compile(r"(.+?)x(.+?) Persona Report (\d{8}) (\d{6})\.xlsx$")
AGE_LABELS = {
    "1825": "18-25岁",
    "2530": "25-30岁",
    "3035": "30-35岁",
    "3540": "35-40岁",
    "4045": "40-45岁",
    "4550": "45-50岁",
    "5055": "50-55岁",
    "55+": "55岁以上",
}


def normalize_city(city: str) -> str:
    return "四五线" if city == "四线五线" else city


def percentile_map(values: dict[str, float]) -> dict[str, float]:
    items = sorted(values.items(), key=lambda item: item[1])
    if len(items) == 1:
        return {items[0][0]: 100.0}
    output: dict[str, float] = {}
    i = 0
    while i < len(items):
        j = i
        while j + 1 < len(items) and items[j + 1][1] == items[i][1]:
            j += 1
        avg_rank = (i + j) / 2
        percentile = 100 * avg_rank / (len(items) - 1)
        for key, _ in items[i : j + 1]:
            output[key] = percentile
        i = j + 1
    return output


def chi_square_2x2(a: int, row_total: int, col_total: int, grand_total: int) -> tuple[float, float]:
    b = row_total - a
    c = col_total - a
    d = grand_total - row_total - c
    denominator = (a + b) * (c + d) * (a + c) * (b + d)
    if denominator <= 0:
        return 0.0, 1.0
    chi2 = grand_total * (a * d - b * c) ** 2 / denominator
    return chi2, math.erfc(math.sqrt(chi2 / 2))


def p_label(value: float | None) -> str:
    if value is None:
        return ""
    if value < 0.001:
        return "p<0.001"
    if value < 0.01:
        return "p<0.01"
    if value < 0.05:
        return "p<0.05"
    return f"p={value:.3f}"


def escape_md(value: Any) -> str:
    return str(value).replace("|", "\\|")


def format_num(value: float | None, digits: int = 1) -> str:
    if value is None:
        return ""
    return f"{value:.{digits}f}"


def load_ingest_helpers(path: Path):
    spec = importlib.util.spec_from_file_location("cid_ingest", path)
    ingest = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(ingest)
    return ingest


class AudienceRelevanceAnalyzer:
    def __init__(self, config: dict[str, Any], source_dir: Path, ingest_path: Path):
        self.config = config
        self.source_dir = source_dir
        self.ingest = load_ingest_helpers(ingest_path)
        self.segments: list[dict[str, Any]] = []
        self.profiles: list[dict[str, Any]] = []
        self.entities: list[dict[str, Any]] = []
        self.buckets: list[dict[str, Any]] = []
        self.segment_by_code: dict[str, dict[str, Any]] = {}
        self.primary_codes: list[str] = []
        self.aux_codes: list[str] = []
        self.primary_grand_total = 0
        self.feature_stats: dict[str, dict[str, Any]] = {}
        self.aux_feature_stats: dict[str, dict[str, Any]] = {}
        self.objective_scores: dict[str, dict[str, dict[str, Any]]] = {}
        self.ranking_rows: list[dict[str, Any]] = []

    def parse_all_workbooks(self) -> None:
        files = sorted(self.source_dir.glob("*.xlsx"))
        if not files:
            raise SystemExit(f"No xlsx files found in {self.source_dir}")
        for path in files:
            segment, profile_rows, entity_rows, bucket_rows = self.parse_workbook(path)
            self.segments.append(segment)
            self.profiles.extend(profile_rows)
            self.entities.extend(entity_rows)
            self.buckets.extend(bucket_rows)

        self.segment_by_code = {segment["segment_code"]: segment for segment in self.segments}
        self.primary_codes = [segment["segment_code"] for segment in self.segments if segment["segment_type"] == "age_city"]
        self.aux_codes = [segment["segment_code"] for segment in self.segments if segment["segment_type"] == "life_stage_city"]
        self.primary_grand_total = sum(int(self.segment_by_code[code]["segment_total_one_id"]) for code in self.primary_codes)
        if len(self.primary_codes) != 40:
            raise SystemExit(f"Expected 40 primary age-city segments, got {len(self.primary_codes)}")

    def parse_workbook(self, path: Path) -> tuple[dict[str, Any], list[dict], list[dict], list[dict]]:
        name = path.name
        age_match = AGE_FILE_RE.match(name)
        life_match = LIFE_FILE_RE.match(name)
        if age_match:
            age_code, raw_city, date_s, time_s = age_match.groups()
            city = normalize_city(raw_city)
            segment_code = f"{age_code}x{city}"
            segment = {
                "segment_code": segment_code,
                "segment_name": f"{AGE_LABELS.get(age_code, age_code)}x{city}",
                "segment_type": "age_city",
                "segment_axis": "age_band",
                "age_band_code": age_code,
                "age_band": AGE_LABELS.get(age_code, age_code),
                "life_stage": "",
                "city_tier": city,
                "city_tier_raw": raw_city,
                "file_name": name,
                "report_generated_at": self.ingest.dt_from_file(date_s, time_s),
            }
        elif life_match:
            life_stage, raw_city, date_s, time_s = life_match.groups()
            city = normalize_city(raw_city)
            segment_code = f"{life_stage}x{city}"
            segment = {
                "segment_code": segment_code,
                "segment_name": f"{life_stage}x{city}",
                "segment_type": "life_stage_city",
                "segment_axis": "life_stage",
                "age_band_code": "",
                "age_band": "",
                "life_stage": life_stage,
                "city_tier": city,
                "city_tier_raw": raw_city,
                "file_name": name,
                "report_generated_at": self.ingest.dt_from_file(date_s, time_s),
            }
        else:
            raise ValueError(f"Unexpected file name: {name}")

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["人群画像"]
        rows = [tuple(self.ingest.clean(v) for v in row) for row in ws.iter_rows(values_only=True)]
        segment["segment_definition"] = self.ingest.text(rows[5][5]) if len(rows) > 5 and len(rows[5]) > 5 else ""

        ecom_header = None
        app_header = None
        for idx, row in enumerate(rows, start=1):
            joined = "|".join(str(v) for v in row if v is not None)
            if ecom_header is None and ("购买频次" in joined or "购买金额" in joined):
                ecom_header = idx
            if app_header is None and ("日均使用频次" in joined or "日均使用时长" in joined):
                app_header = idx
        if ecom_header is None:
            raise ValueError(f"No purchase header found: {name}")

        total = None
        for row in rows[7 : ecom_header - 1]:
            vals = list(row) + [None] * 8
            one_id, ratio = vals[5], vals[6]
            if isinstance(one_id, (int, float)) and isinstance(ratio, (int, float)) and abs(ratio - 1) < 1e-9:
                total = max(total or 0, int(one_id))
        segment["segment_total_one_id"] = total or 0

        profile_rows = self.ingest.parse_profile_rows(rows, ecom_header, segment_code)
        entity_rows, bucket_rows = self.ingest.parse_purchase_rows(rows, ecom_header, app_header, segment_code)
        entity_rows.extend(self.ingest.parse_media_rows(rows, app_header, segment_code) if app_header else [])
        return segment, profile_rows, entity_rows, bucket_rows

    def build_features(self) -> None:
        for key, feature in self.config["features"].get("purchase_categories", {}).items():
            counts = self.build_purchase_category_counts(self.primary_codes, feature["entity_name"])
            self.add_count_feature(key, feature["label"], feature.get("family_label", "品类购买"), counts)

        for key, feature in self.config["features"].get("profile", {}).items():
            rows = [tuple(row) for row in feature["rows"]]
            counts = self.build_profile_counts(self.primary_codes, rows)
            self.add_count_feature(key, feature["label"], feature.get("family_label", "画像证据"), counts)

        for key, feature in self.config["features"].get("media_categories", {}).items():
            self.add_media_feature(key, feature["label"], feature["category"])

    def build_purchase_category_counts(self, codes: list[str], entity_name: str) -> dict[str, int]:
        counts = {code: 0 for code in codes}
        for row in self.entities:
            if row["segment_code"] not in counts:
                continue
            if row["domain"] == "purchase" and row["entity_type"] == "category" and row["entity_name"] == entity_name:
                counts[row["segment_code"]] += int(row["one_id"])
        return counts

    def build_profile_counts(self, codes: list[str], rows: list[tuple[str, str, str]]) -> dict[str, int]:
        wanted = set(rows)
        counts = {code: 0 for code in codes}
        for row in self.profiles:
            if row["segment_code"] not in counts:
                continue
            row_key = (row["level1_category"], row["level2_category"], row["level3_value"])
            if row_key in wanted:
                counts[row["segment_code"]] += int(row["one_id"])
        return counts

    def add_count_feature(
        self,
        key: str,
        label: str,
        family_label: str,
        counts: dict[str, int],
        codes: list[str] | None = None,
        universe_total: int | None = None,
        target: dict[str, dict[str, Any]] | None = None,
    ) -> None:
        codes = codes or self.primary_codes
        universe_total = universe_total or self.primary_grand_total
        target = target if target is not None else self.feature_stats
        universe_count = sum(counts.values())
        universe_rate = universe_count / universe_total if universe_total else 0
        affinity: dict[str, float] = {}
        ratios: dict[str, float] = {}
        chi2: dict[str, float] = {}
        p_values: dict[str, float] = {}
        for code in codes:
            segment_total = int(self.segment_by_code[code]["segment_total_one_id"])
            count = int(counts.get(code, 0))
            ratio = count / segment_total if segment_total else 0
            ratios[code] = ratio
            affinity[code] = ratio / universe_rate * 100 if universe_rate else 0
            chi2[code], p_values[code] = chi_square_2x2(count, segment_total, universe_count, universe_total)
        target[key] = {
            "kind": "observed_metric",
            "label": label,
            "family_label": family_label,
            "counts": counts,
            "ratios": ratios,
            "universe_count": universe_count,
            "universe_rate": universe_rate,
            "affinity": affinity,
            "chi2": chi2,
            "p_values": p_values,
            "percentile": percentile_map(affinity),
        }

    def add_media_feature(self, key: str, label: str, category: str) -> None:
        app_counts: dict[str, dict[str, int]] = {}
        primary_set = set(self.primary_codes)
        for row in self.entities:
            if row["segment_code"] not in primary_set:
                continue
            if row["domain"] != "media" or row["level2_category"] != category:
                continue
            app_counts.setdefault(row["entity_name"], {code: 0 for code in self.primary_codes})
            app_counts[row["entity_name"]][row["segment_code"]] += int(row["one_id"])

        app_affinity: dict[str, dict[str, float]] = {}
        app_weights: dict[str, float] = {}
        for app, counts in app_counts.items():
            universe_count = sum(counts.values())
            if universe_count <= 0:
                continue
            universe_rate = universe_count / self.primary_grand_total
            app_weights[app] = math.sqrt(universe_count)
            app_affinity[app] = {}
            for code in self.primary_codes:
                segment_total = int(self.segment_by_code[code]["segment_total_one_id"])
                ratio = counts[code] / segment_total if segment_total else 0
                app_affinity[app][code] = ratio / universe_rate * 100 if universe_rate else 0

        weight_total = sum(app_weights.values())
        derived_affinity: dict[str, float] = {}
        top_apps: dict[str, list[tuple[str, float]]] = {}
        for code in self.primary_codes:
            derived_affinity[code] = (
                sum(app_affinity[app][code] * weight for app, weight in app_weights.items()) / weight_total
                if weight_total
                else 0
            )
            top_apps[code] = sorted(
                ((app, app_affinity[app][code]) for app in app_affinity),
                key=lambda item: item[1],
                reverse=True,
            )[:3]

        self.feature_stats[key] = {
            "kind": "derived_media_category",
            "label": label,
            "family_label": "媒体触达",
            "category": category,
            "affinity": derived_affinity,
            "percentile": percentile_map(derived_affinity),
            "top_apps": top_apps,
        }

    def score_objectives(self) -> None:
        self.objective_scores = {key: self.score_objective(key) for key in self.config["objectives"]}

    def score_objective(self, objective_key: str) -> dict[str, dict[str, Any]]:
        objective = self.config["objectives"][objective_key]
        scores: dict[str, dict[str, Any]] = {}
        for code in self.primary_codes:
            family_scores = {}
            for family, keys in objective["families"].items():
                family_scores[family] = statistics.mean(self.feature_stats[key]["percentile"][code] for key in keys)
            scores[code] = {"score": statistics.mean(family_scores.values()), "families": family_scores}
        ranked_codes = sorted(self.primary_codes, key=lambda code: scores[code]["score"], reverse=True)
        for rank, code in enumerate(ranked_codes, start=1):
            scores[code]["rank"] = rank
        return scores

    def evidence_summary(self, objective_key: str, code: str) -> str:
        objective = self.config["objectives"][objective_key]
        parts = []
        for family, keys in objective["families"].items():
            best_key = max(keys, key=lambda item: self.feature_stats[item]["percentile"][code])
            stat = self.feature_stats[best_key]
            if stat["kind"] == "derived_media_category":
                app_text = "、".join(f"{app} AI={aff:.1f}" for app, aff in stat["top_apps"][code][:2])
                parts.append(f"{family}: {stat['label']} AI={stat['affinity'][code]:.1f} ({app_text})")
            else:
                parts.append(
                    f"{family}: {stat['label']} AI={stat['affinity'][code]:.1f}, "
                    f"{p_label(stat['p_values'][code])}"
                )
        return "；".join(parts)

    def plain_feature_phrase(self, key: str, code: str) -> str:
        stat = self.feature_stats[key]
        affinity = stat["affinity"][code]
        p_value = stat.get("p_values", {}).get(code)
        weak = affinity < 100 or (p_value is not None and p_value >= 0.05)
        if not weak:
            return self.config.get("plain_phrases", {}).get(key, f"{stat['label']}明显高于整体")
        if stat["kind"] == "derived_media_category":
            return f"{stat['label']}触达没有明显高于整体，说明这个渠道不是该人群的主要优势"
        return f"{stat['label']}没有明显高于整体，说明这个维度不是该人群的主要优势"

    def plain_explanation(self, objective_key: str, code: str) -> str:
        score = self.objective_scores[objective_key][code]["score"]
        objective = self.config["objectives"][objective_key]
        best_keys = [max(keys, key=lambda item: self.feature_stats[item]["percentile"][code]) for keys in objective["families"].values()]
        reasons = [self.plain_feature_phrase(key, code) for key in best_keys]
        if objective_key == "conversion":
            objective_text = "转化匹配度"
        else:
            objective_text = "品牌心智匹配度"
        if score >= 70:
            level = "高"
        elif score >= 55:
            level = "较高"
        elif score >= 40:
            level = "中等"
        else:
            level = "偏低"
        if score >= 55:
            prefix = f"这个人群和{self.config['input_name']}的{objective_text}{level}，主要看几个方面："
            suffix = "统计结果也显示，这些差异比较明显，不太像是偶然出现的。"
        elif score >= 40:
            prefix = f"这个人群和{self.config['input_name']}有一定{objective_text}，但优势不算特别突出，主要看几个方面："
            suffix = "所以它可以作为补充人群看待，但不应放在第一优先级。"
        else:
            prefix = f"这个人群的{objective_text}偏低，当前数据看不出特别强的优势；相对能看到的情况是："
            suffix = "整体来看，它更适合放在低优先级或长尾覆盖里。"
        close = self.config.get("plain_closing", {}).get(objective_key, "")
        return prefix + "；".join(reasons) + "。" + close + suffix

    def build_auxiliary_stats(self) -> None:
        aux_config = self.config.get("auxiliary", {})
        features = aux_config.get("features", [])
        aux_total = sum(int(self.segment_by_code[code]["segment_total_one_id"]) for code in self.aux_codes)
        for key in features:
            if key in self.config["features"].get("purchase_categories", {}):
                feature = self.config["features"]["purchase_categories"][key]
                counts = self.build_purchase_category_counts(self.aux_codes, feature["entity_name"])
                self.add_count_feature(key, feature["label"], feature.get("family_label", "辅助品类"), counts, self.aux_codes, aux_total, self.aux_feature_stats)
            elif key in self.config["features"].get("profile", {}):
                feature = self.config["features"]["profile"][key]
                rows = [tuple(row) for row in feature["rows"]]
                counts = self.build_profile_counts(self.aux_codes, rows)
                self.add_count_feature(key, feature["label"], feature.get("family_label", "辅助画像"), counts, self.aux_codes, aux_total, self.aux_feature_stats)

    def auxiliary_summary(self, city_tier: str) -> str:
        if not self.aux_feature_stats:
            return ""
        features = list(self.aux_feature_stats)
        top_n = int(self.config.get("auxiliary", {}).get("top_n", 2))
        candidates = [code for code in self.aux_codes if self.segment_by_code[code]["city_tier"] == city_tier]
        scored = []
        for code in candidates:
            score = statistics.mean(self.aux_feature_stats[key]["percentile"][code] for key in features)
            scored.append((score, code))
        parts = []
        for _, code in sorted(scored, reverse=True)[:top_n]:
            segment = self.segment_by_code[code]
            evidence = "、".join(f"{self.aux_feature_stats[key]['label']}AI={self.aux_feature_stats[key]['affinity'][code]:.1f}" for key in features)
            parts.append(f"{segment['segment_name']} {evidence}")
        return "；".join(parts)

    def build_ranking_rows(self) -> None:
        rows = []
        for code in self.primary_codes:
            conversion = self.objective_scores["conversion"][code]
            brand_mind = self.objective_scores["brand_mind"][code]
            segment = self.segment_by_code[code]
            row = {
                "audience": segment["segment_name"],
                "segment_code": code,
                "age_band": segment["age_band"],
                "city_tier": segment["city_tier"],
                "segment_total_one_id": segment["segment_total_one_id"],
                "match_status": self.config.get("match_status", ""),
                "conversion_rank": conversion["rank"],
                "conversion_score": round(conversion["score"], 1),
                "conversion_evidence": self.evidence_summary("conversion", code),
                "conversion_plain_explanation": self.plain_explanation("conversion", code),
                "brand_mind_rank": brand_mind["rank"],
                "brand_mind_score": round(brand_mind["score"], 1),
                "brand_mind_evidence": self.evidence_summary("brand_mind", code),
                "brand_mind_plain_explanation": self.plain_explanation("brand_mind", code),
                "auxiliary_life_stage_evidence": self.auxiliary_summary(segment["city_tier"]),
            }
            for family, value in conversion["families"].items():
                row[f"conversion_{family}_score"] = round(value, 1)
            for family, value in brand_mind["families"].items():
                row[f"brand_mind_{family}_score"] = round(value, 1)
            rows.append(row)
        self.ranking_rows = sorted(rows, key=lambda row: int(row["conversion_rank"]))

    def direct_counts(self) -> dict[str, int]:
        counts = {term: 0 for term in self.config.get("direct_match_terms", [])}
        strict_terms = self.config.get("strict_observed_terms", [])
        for item in strict_terms:
            counts[item["label"]] = 0
        for row in self.entities:
            hay = "|".join([row["entity_name"], row["label_path"], row["raw_label"]])
            for term in self.config.get("direct_match_terms", []):
                if term in hay:
                    counts[term] += 1
            for item in strict_terms:
                if (
                    row["domain"] == item.get("domain")
                    and row["entity_type"] == item.get("entity_type")
                    and row["entity_name"] == item.get("entity_name")
                ):
                    counts[item["label"]] += 1
        return counts

    def write_outputs(self, out_dir: Path) -> dict[str, Path]:
        out_dir.mkdir(parents=True, exist_ok=True)
        prefix = f"brand_product_audience_{self.config['analysis_id']}"
        ranking_path = out_dir / f"{prefix}_ranking.csv"
        evidence_path = out_dir / f"{prefix}_evidence.csv"
        report_path = out_dir / f"{prefix}_report.md"

        with ranking_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(self.ranking_rows[0].keys()))
            writer.writeheader()
            writer.writerows(self.ranking_rows)

        self.write_evidence(evidence_path)
        report_path.write_text(self.render_report(), encoding="utf-8")
        return {"ranking": ranking_path, "evidence": evidence_path, "report": report_path}

    def write_evidence(self, path: Path) -> None:
        fieldnames = [
            "objective",
            "audience",
            "segment_code",
            "age_band",
            "city_tier",
            "feature_key",
            "feature_label",
            "feature_family",
            "feature_kind",
            "one_id",
            "segment_total_one_id",
            "ratio",
            "universe_rate",
            "affinity_index",
            "chi_square",
            "p_value",
            "percentile_score",
        ]
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for objective_key, objective in self.config["objectives"].items():
                objective_features = {key for keys in objective["families"].values() for key in keys}
                for code in self.primary_codes:
                    segment = self.segment_by_code[code]
                    for key in sorted(objective_features):
                        stat = self.feature_stats[key]
                        row = {
                            "objective": objective_key,
                            "audience": segment["segment_name"],
                            "segment_code": code,
                            "age_band": segment["age_band"],
                            "city_tier": segment["city_tier"],
                            "feature_key": key,
                            "feature_label": stat["label"],
                            "feature_family": stat["family_label"],
                            "feature_kind": stat["kind"],
                            "segment_total_one_id": segment["segment_total_one_id"],
                            "affinity_index": f"{stat['affinity'][code]:.4f}",
                            "percentile_score": f"{stat['percentile'][code]:.4f}",
                        }
                        if stat["kind"] == "derived_media_category":
                            row.update({"one_id": "", "ratio": "", "universe_rate": "", "chi_square": "", "p_value": ""})
                        else:
                            row.update(
                                {
                                    "one_id": stat["counts"][code],
                                    "ratio": f"{stat['ratios'][code]:.8f}",
                                    "universe_rate": f"{stat['universe_rate']:.8f}",
                                    "chi_square": f"{stat['chi2'][code]:.4f}",
                                    "p_value": f"{stat['p_values'][code]:.8g}",
                                }
                            )
                        writer.writerow(row)

    def render_top_table(self, objective_key: str, n: int = 10) -> str:
        rank_field = f"{objective_key}_rank"
        score_field = f"{objective_key}_score"
        evidence_field = f"{objective_key}_evidence"
        plain_field = f"{objective_key}_plain_explanation"
        rows = sorted(self.ranking_rows, key=lambda row: int(row[rank_field]))[:n]
        label = self.config["objectives"][objective_key]["label"]
        lines = [f"| {label} 排名 | 人群 | Score | 核心证据 | 直白解释 |", "|---:|---|---:|---|---|"]
        for row in rows:
            lines.append(
                f"| {row[rank_field]} | {escape_md(row['audience'])} | "
                f"{format_num(float(row[score_field]), 1)} | {escape_md(row[evidence_field])} | "
                f"{escape_md(row[plain_field])} |"
            )
        return "\n".join(lines)

    def render_ranking_table(self) -> str:
        lines = [
            "| 转化排名 | 人群 | 转化分 | Brand Mind 排名 | Brand Mind 分 | 转化证据摘要 | 品牌心智证据摘要 | 辅助人生阶段证据 |",
            "|---:|---|---:|---:|---:|---|---|---|",
        ]
        for row in self.ranking_rows:
            lines.append(
                "| "
                + " | ".join(
                    [
                        str(row["conversion_rank"]),
                        escape_md(row["audience"]),
                        format_num(float(row["conversion_score"]), 1),
                        str(row["brand_mind_rank"]),
                        format_num(float(row["brand_mind_score"]), 1),
                        escape_md(row["conversion_evidence"]),
                        escape_md(row["brand_mind_evidence"]),
                        escape_md(row["auxiliary_life_stage_evidence"]),
                    ]
                )
                + " |"
            )
        return "\n".join(lines)

    def render_report(self) -> str:
        direct_counts = self.direct_counts()
        source_lines = ["| 来源 | 路径 / URL | 抓取或读取日期 | 用途 |", "|---|---|---|---|"]
        for source in self.config.get("sources", []):
            source_lines.append(f"| {escape_md(source['name'])} | {escape_md(source['url'])} | {source['date']} | {escape_md(source['usage'])} |")

        attr_lines = ["| 字段 | 结果 |", "|---|---|"]
        for key, value in self.config.get("product_attributes", {}).items():
            attr_lines.append(f"| {escape_md(key)} | {escape_md(value)} |")

        direct_lines = ["| 关键词 | 命中行数 |", "|---|---:|"]
        for key, value in direct_counts.items():
            direct_lines.append(f"| {escape_md(key)} | {value} |")

        def feature_label(key: str) -> str:
            stat = self.feature_stats.get(key)
            return stat["label"] if stat else key

        conversion_families = "\n".join(
            f"- {family}：{'、'.join(feature_label(key) for key in keys)}。"
            for family, keys in self.config["objectives"]["conversion"]["families"].items()
        )
        brand_families = "\n".join(
            f"- {family}：{'、'.join(feature_label(key) for key in keys)}。"
            for family, keys in self.config["objectives"]["brand_mind"]["families"].items()
        )

        return f"""# {self.config.get('report_title', self.config['product_name'] + ' 人群相关性评估')}

更新时间：{self.config.get('report_date', '')}

## 结论口径

- 输入：{self.config.get('input_name', '')}。
- 标准化对象：{self.config.get('product_name', '')}。
- 主分析人群：40 个 `年龄段 x 城市级别` 切片。
- 辅助数据源：30 个 `人生阶段 x 城市级别` 切片，仅用于解释，不参与主排序分数。
- 匹配状态：`{self.config.get('match_status', '')}`。
- Score 是 40 个年龄城市人群内的相对证据强度，范围 0-100，不是实际转化率或品牌认知率。

## 来源审计

{chr(10).join(source_lines)}

## 产品属性归一化

{chr(10).join(attr_lines)}

## CID 直接命中检查

{chr(10).join(direct_lines)}

## 评分方法

### Conversion / 效果转化

{conversion_families}

### Brand Mind / 品牌心智

{brand_families}

## Top 10

### Conversion

{self.render_top_table('conversion', 10)}

### Brand Mind

{self.render_top_table('brand_mind', 10)}

## 40 个年龄城市人群并排排序

{self.render_ranking_table()}

## 解读边界

- 如果是 `category_observed`，代表使用了可严格观测的品类/媒体实体，不是品牌或产品本身的直接购买观测。
- 由于样本量很大，卡方检验容易显著；排序主要看 Affinity Index 的方向和相对强度，卡方用于确认差异并非随机波动。
- 当前没有用户级原始数据，因此不做多变量归因，也不声称年龄城市切片和人生阶段切片之间存在严格因果关系。
- 如果后续拿到品牌、产品或更细品类的直接数据，应优先替换本结果。
"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", required=True)
    parser.add_argument("--source-dir", default="/Users/apple/Downloads/User/CID 人群")
    parser.add_argument("--out-dir", default="outputs/brand_product_audience")
    default_ingest = Path(__file__).with_name("ingest_cid_persona.py")
    parser.add_argument("--ingest-script", default=str(default_ingest))
    args = parser.parse_args()

    config = json.loads(Path(args.config).read_text(encoding="utf-8"))
    analyzer = AudienceRelevanceAnalyzer(config, Path(args.source_dir), Path(args.ingest_script))
    analyzer.parse_all_workbooks()
    analyzer.build_features()
    analyzer.score_objectives()
    analyzer.build_auxiliary_stats()
    analyzer.build_ranking_rows()
    paths = analyzer.write_outputs(Path(args.out_dir))
    print(f"primary_segments={len(analyzer.primary_codes)} auxiliary_segments={len(analyzer.aux_codes)}")
    for name, path in paths.items():
        print(f"{name}={path}")


if __name__ == "__main__":
    main()
